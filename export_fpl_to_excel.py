import os
import csv
import json
import time
import shutil
from datetime import datetime
import urllib3
import requests

# ปิด SSL warning
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
}

CHIP_DISPLAY = {
    "3xc": "Triple Captain (TC)",
    "bboost": "Bench Boost (BB)",
    "wildcard": "Wildcard (WC)",
    "freehit": "Free Hit (FH)"
}

MONTH_THAI = {
    1: "มกราคม (Jan)", 2: "กุมภาพันธ์ (Feb)", 3: "มีนาคม (Mar)",
    4: "เมษายน (Apr)", 5: "พฤษภาคม (May)", 6: "มิถุนายน (Jun)",
    7: "กรกฎาคม (Jul)", 8: "สิงหาคม (Aug)", 9: "กันยายน (Sep)",
    10: "ตุลาคม (Oct)", 11: "พฤศจิกายน (Nov)", 12: "ธันวาคม (Dec)"
}

# แถบเมนูด้านบนพร้อมปุ่มย้อนกลับไปหน้าหลัก (index.html) + ปุ่มเซฟภาพ
NAV_ACTION_BAR_HTML = """
<div class="action-bar">
    <a href="index.html" class="btn-back">⬅ กลับหน้าหลัก (Home)</a>
    <button onclick="downloadAsImage()" class="btn-dl">📸 บันทึกเป็นรูปภาพ (PNG)</button>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<script>
function downloadAsImage() {
    const btn = document.querySelector('.action-bar');
    btn.style.display = 'none';
    const target = document.querySelector('.container');
    html2canvas(target, {
        scale: 2,
        backgroundColor: '#14001a',
        useCORS: true
    }).then(canvas => {
        const link = document.createElement('a');
        link.download = document.title.replace(/[^a-zA-Z0-9ก-๙_-]/g, '_') + '.png';
        link.href = canvas.toDataURL('image/png');
        link.click();
        btn.style.display = 'flex';
    }).catch(err => {
        btn.style.display = 'flex';
        alert('เกิดข้อผิดพลาดในการสร้างรูปภาพ กรุณาใช้ปุ่ม Print Screen หรือแคปหน้าจอแทนครับ');
    });
}
</script>
"""

ACTION_BAR_CSS = """
.action-bar {
    display: flex;
    justify-content: center;
    align-items: center;
    gap: 12px;
    margin-bottom: 20px;
    flex-wrap: wrap;
}
.btn-back {
    background: rgba(255, 255, 255, 0.08);
    border: 1px solid rgba(255, 255, 255, 0.2);
    color: #ffffff;
    padding: 10px 22px;
    border-radius: 30px;
    font-size: 13.5px;
    font-weight: 700;
    text-decoration: none;
    display: inline-flex;
    align-items: center;
    gap: 6px;
    transition: all 0.2s;
    font-family: 'Kanit', sans-serif;
}
.btn-back:hover {
    background: rgba(255, 255, 255, 0.18);
    border-color: #00ff87;
    color: #00ff87;
    transform: translateY(-2px);
    box-shadow: 0 0 15px rgba(0, 255, 135, 0.3);
}
.btn-dl {
    background: linear-gradient(90deg, #00ff87, #04f5ff);
    color: #120015;
    border: none;
    padding: 11px 24px;
    border-radius: 30px;
    font-size: 13.5px;
    font-weight: 800;
    cursor: pointer;
    font-family: 'Kanit', sans-serif;
    box-shadow: 0 0 20px rgba(0, 255, 135, 0.4);
    transition: transform 0.2s, box-shadow 0.2s;
}
.btn-dl:hover {
    transform: translateY(-2px);
    box-shadow: 0 0 30px rgba(0, 255, 135, 0.7);
}
"""

def load_config():
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    default_config = {
        "league_id": 506845,
        "rules": {
            "no_chip_gameweeks": [1]
        },
        "tie_breaker": {
            "enabled": True,
            "order": [
                "1) คะแนนกัปตันสูงกว่า",
                "2) Bench รวมสูงกว่า",
                "3) จับฉลาก"
            ],
            "description": "กรณีแต้ม GW เท่ากัน: 1) คะแนนกัปตันสูงกว่า 2) Bench รวมสูงกว่า 3) จับฉลาก"
        },
        "prizes": {
            "special_weekly_prizes": {
                "1": "🎁 รางวัลเปิดฤดูกาล GW1 (ห้ามใช้ชิป)",
                "10": "🎁 รางวัลพิเศษ GW10",
                "19": "🎁 รางวัลแชมป์ครึ่งฤดูกาลแรก (Mid-Season)",
                "38": "🎁 รางวัลปิดฤดูกาล GW38"
            },
            "monthly_top_places": 3
        }
    }
    if not os.path.exists(config_path):
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(default_config, f, ensure_ascii=False, indent=2)
        return default_config
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default_config

def get_bootstrap_data():
    url = "https://fantasy.premierleague.com/api/bootstrap-static/"
    res = requests.get(url, headers=HEADERS, verify=False)
    if res.status_code != 200:
        raise Exception(f"Failed to fetch bootstrap: {res.status_code}")
    
    data = res.json()
    events = data.get("events", [])
    player_map = {p["id"]: p["web_name"] for p in data.get("elements", [])}
    
    current_gw = 1
    months_gw_map = {}
    gw_to_month = {}
    
    for ev in events:
        gw_id = ev.get("id")
        if ev.get("is_current"):
            current_gw = gw_id
            
        deadline = ev.get("deadline_time")
        if deadline:
            dt = datetime.strptime(deadline[:10], "%Y-%m-%d")
            m_name = f"{MONTH_THAI.get(dt.month, dt.strftime('%B'))} {dt.year}"
            if m_name not in months_gw_map:
                months_gw_map[m_name] = []
            months_gw_map[m_name].append(gw_id)
            gw_to_month[gw_id] = m_name

    if current_gw == 1:
        for ev in reversed(events):
            if ev.get("finished"):
                current_gw = ev.get("id")
                break
                
    return events, current_gw, player_map, months_gw_map, gw_to_month

def get_live_points_for_gw(gw):
    live_map = {}
    try:
        url = f"https://fantasy.premierleague.com/api/event/{gw}/live/"
        res = requests.get(url, headers=HEADERS, verify=False)
        if res.status_code == 200:
            for el in res.json().get("elements", []):
                live_map[el["id"]] = el.get("stats", {}).get("total_points", 0)
    except Exception:
        pass
    return live_map

def save_excel_safe(sheets_dict, target_path):
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment

    def write_workbook(path):
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                for col_idx, col in enumerate(ws.columns, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 11)
        return path

    try:
        return write_workbook(target_path)
    except PermissionError:
        timestamp = datetime.now().strftime("%H%M%S")
        dir_name, base_name = os.path.split(target_path)
        name, ext = os.path.splitext(base_name)
        new_path = os.path.join(dir_name, f"{name}_{timestamp}{ext}")
        return write_workbook(new_path)

def generate_weekly_standings_html(league_name, league_id, current_gw, display_rows, output_path):
    highest_gw_pts = max((r.get("gw_pts", 0) for r in display_rows), default=0)
    top_gw_scorers = [r for r in display_rows if r.get("gw_pts") == highest_gw_pts]
    top1 = display_rows[0] if len(display_rows) > 0 else None

    table_rows_html = ""
    for r in display_rows:
        rank = r["rank"]
        change = r["rank_change"]
        if change > 0:
            change_badge = f'<span class="badge-up">▲ +{change}</span>'
        elif change < 0:
            change_badge = f'<span class="badge-down">▼ {change}</span>'
        else:
            change_badge = '<span class="badge-same">▬ 0</span>'

        rank_class = "rank-gold" if rank == 1 else ("rank-silver" if rank == 2 else ("rank-bronze" if rank == 3 else ""))
        rank_icon = "🥇 1" if rank == 1 else ("🥈 2" if rank == 2 else ("🥉 3" if rank == 3 else f"#{rank}"))

        chip_badge = f'<span class="chip-tag">{r["chip"]}</span>' if r.get("chip") and r.get("chip") != "-" else ""
        gw_high_class = "gw-high" if r.get("gw_pts") == highest_gw_pts and highest_gw_pts > 0 else ""

        # แท็กข้อมูลกัปตันและสำรอง
        cap_badge = f'<span class="cap-pill">© {r["captain_name"]} ({r["captain_pts"]} pts)</span>' if r.get("captain_name") and r.get("captain_name") != "-" else ""
        bench_badge = f'<span class="bench-pill">🪑 สำรอง: {r["bench_pts"]} pts</span>' if "bench_pts" in r else ""

        table_rows_html += f"""
        <tr class="{rank_class}">
            <td class="col-rank">{rank_icon}</td>
            <td class="col-change">{change_badge}</td>
            <td class="col-team">
                <div class="team-title">{r['team_name']}</div>
                <div class="manager-title">
                    {r['player_name']} {chip_badge}
                </div>
                <div class="meta-row">
                    {cap_badge}
                    {bench_badge}
                </div>
            </td>
            <td class="col-gw {gw_high_class}">+{r['gw_pts']}</td>
            <td class="col-total">{r['total_pts']}</td>
        </tr>
        """

    html = f"""<!DOCTYPE html>
<html lang="th">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Weekly Standings GW{current_gw} - {league_name}</title>
    <link href="https://fonts.googleapis.com/css2?family=Kanit:wght@300;400;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root {{ --pl-purple: #37003c; --pl-green: #00ff87; --pl-pink: #e90052; --pl-cyan: #04f5ff; --pl-gold: #ffd700; }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ background: linear-gradient(135deg, #1b001e 0%, #0d000f 100%); font-family: 'Kanit', sans-serif; color: #fff; padding: 25px 15px; display: flex; flex-direction: column; align-items: center; }}
        {ACTION_BAR_CSS}
        .container {{ max-width: 920px; width: 100%; background: rgba(36, 4, 40, 0.95); backdrop-filter: blur(15px); border: 1px solid rgba(255, 255, 255, 0.1); border-radius: 24px; padding: 35px 25px; box-shadow: 0 25px 60px rgba(0, 0, 0, 0.6), 0 0 30px rgba(0, 255, 135, 0.15); }}
        .header {{ text-align: center; margin-bottom: 25px; }}
        .badge-type {{ display: inline-block; background: linear-gradient(90deg, var(--pl-pink), #963cff); color: white; padding: 5px 16px; border-radius: 30px; font-size: 12px; font-weight: 700; letter-spacing: 1.5px; text-transform: uppercase; margin-bottom: 10px; }}
        .title {{ font-size: 30px; font-weight: 800; background: linear-gradient(90deg, #ffffff, #00ff87); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-bottom: 5px; }}
        .subtitle {{ color: #c4a9d4; font-size: 14.5px; }}
        .highlights-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); gap: 15px; margin-bottom: 25px; }}
        .card-hl {{ background: rgba(255, 255, 255, 0.04); border: 1px solid rgba(255, 255, 255, 0.08); border-radius: 16px; padding: 16px; display: flex; align-items: center; gap: 15px; }}
        .card-hl.gold-hl {{ border-left: 4px solid var(--pl-gold); background: rgba(255, 215, 0, 0.06); }}
        .card-hl.green-hl {{ border-left: 4px solid var(--pl-green); background: rgba(0, 255, 135, 0.06); }}
        .hl-icon {{ font-size: 32px; }}
        .hl-label {{ font-size: 11px; color: #a98fb9; text-transform: uppercase; font-weight: 700; }}
        .hl-name {{ font-size: 15px; font-weight: 700; color: #fff; margin: 2px 0; }}
        .hl-pts {{ font-size: 13.5px; color: var(--pl-green); font-weight: 700; }}
        .table-wrap {{ overflow-x: auto; border-radius: 16px; border: 1px solid rgba(255, 255, 255, 0.07); }}
        table {{ width: 100%; border-collapse: collapse; text-align: left; }}
        thead {{ background: linear-gradient(90deg, #2c0032, #1b001f); }}
        th {{ padding: 14px 16px; font-size: 12.5px; font-weight: 700; color: #d1b4e3; letter-spacing: 0.8px; text-transform: uppercase; }}
        td {{ padding: 13px 16px; border-bottom: 1px solid rgba(255, 255, 255, 0.05); font-size: 13.5px; }}
        tr:hover {{ background: rgba(255, 255, 255, 0.04); }}
        .col-rank {{ font-weight: 800; font-size: 15px; width: 65px; text-align: center; }}
        .rank-gold {{ background: rgba(255, 215, 0, 0.07); }} .rank-gold .col-rank {{ color: var(--pl-gold); }}
        .rank-silver {{ background: rgba(220, 220, 220, 0.05); }} .rank-silver .col-rank {{ color: #e0e0e0; }}
        .rank-bronze {{ background: rgba(205, 127, 50, 0.05); }} .rank-bronze .col-rank {{ color: #cd7f32; }}
        .col-change {{ width: 75px; text-align: center; }}
        .badge-up {{ background: rgba(0, 255, 135, 0.15); color: var(--pl-green); padding: 3px 8px; border-radius: 12px; font-size: 11px; font-weight: 700; }}
        .badge-down {{ background: rgba(233, 0, 82, 0.15); color: var(--pl-pink); padding: 3px 8px; border-radius: 12px; font-size: 11px; font-weight: 700; }}
        .badge-same {{ color: #7b628a; font-size: 11px; font-weight: 600; }}
        .col-team {{ width: 50%; }}
        .team-title {{ font-weight: 700; font-size: 14.5px; color: #ffffff; }}
        .manager-title {{ font-size: 11.5px; color: #ad94be; margin-top: 2px; }}
        .chip-tag {{ background: #963cff; color: #fff; padding: 1px 7px; border-radius: 8px; font-size: 10px; font-weight: 700; margin-left: 6px; }}
        .meta-row {{ margin-top: 5px; display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }}
        .cap-pill {{ background: rgba(4, 245, 255, 0.12); border: 1px solid rgba(4, 245, 255, 0.3); color: #04f5ff; font-size: 11px; font-weight: 700; padding: 2px 8px; border-radius: 10px; }}
        .bench-pill {{ background: rgba(255, 215, 0, 0.1); border: 1px solid rgba(255, 215, 0, 0.25); color: #ffd700; font-size: 11px; font-weight: 600; padding: 2px 8px; border-radius: 10px; }}
        .col-gw {{ font-weight: 700; text-align: center; width: 85px; color: #e0d0eb; }}
        .col-gw.gw-high {{ color: var(--pl-green); font-weight: 800; text-shadow: 0 0 10px rgba(0, 255, 135, 0.4); }}
        .col-total {{ font-weight: 800; font-size: 15.5px; text-align: center; width: 85px; color: var(--pl-cyan); }}
        .footer {{ text-align: center; margin-top: 20px; color: #8b6e9c; font-size: 11.5px; }}
    </style>
</head>
<body>
    {NAV_ACTION_BAR_HTML}
    <div class="container">
        <div class="header">
            <div class="badge-type">📊 WEEKLY LEADERBOARD | LEAGUE {league_id}</div>
            <h1 class="title">{league_name}</h1>
            <p class="subtitle">ตารางคะแนนรวมและอันดับล่าสุดประจำ <strong>GAMEWEEK {current_gw}</strong></p>
        </div>
        <div class="highlights-grid">
            <div class="card-hl gold-hl">
                <div class="hl-icon">👑</div>
                <div>
                    <div class="hl-label">จ่าฝูง (League Leader)</div>
                    <div class="hl-name">{top1['team_name'] if top1 else '-'}</div>
                    <div class="hl-pts">{top1['total_pts'] if top1 else 0} คะแนนรวม</div>
                </div>
            </div>
            <div class="card-hl green-hl">
                <div class="hl-icon">🔥</div>
                <div>
                    <div class="hl-label">แต้มสูงสุด GW{current_gw}</div>
                    <div class="hl-name">{top_gw_scorers[0]['team_name'] if top_gw_scorers else '-'}</div>
                    <div class="hl-pts">+{highest_gw_pts} แต้มสัปดาห์นี้</div>
                </div>
            </div>
        </div>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th class="col-rank">อันดับ</th>
                        <th class="col-change">ขยับ</th>
                        <th class="col-team">ชื่อทีม, ผู้จัดการทีม & กัปตัน</th>
                        <th class="col-gw">แต้ม GW{current_gw}</th>
                        <th class="col-total">คะแนนรวม</th>
                    </tr>
                </thead>
                <tbody>{table_rows_html}</tbody>
            </table>
        </div>
        <div class="footer">Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')} | FPL Auto Exporter</div>
    </div>
</body>
</html>
"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return output_path

def generate_monthly_awards_html(league_name, league_id, current_gw, monthly_summary_data, top_n_places, output_path):
    month_sections_html = ""

    for month_title, data in monthly_summary_data.items():
        if not data["rows"]:
            continue
        
        gw_range_str = f"GW{min(data['gws'])}-{max(data['gws'])}"
        is_active = current_gw in data["gws"]
        status_tag = '<span class="m-status active">กำลังแข่งขัน 🔥</span>' if is_active else '<span class="m-status done">จบเดือนแล้ว 🏁</span>'

        rows = data["rows"]
        m_top1 = rows[0] if len(rows) > 0 else None
        m_top2 = rows[1] if len(rows) > 1 else None
        m_top3 = rows[2] if len(rows) > 2 else None

        podium_html = f"""
        <div class="podium-grid">
            <div class="p-card p-gold">
                <div class="p-badge">🥇 แชมป์ประจำเดือน #{1}</div>
                <div class="p-team">{m_top1['team_name'] if m_top1 else '-'}</div>
                <div class="p-mgr">{m_top1['player_name'] if m_top1 else '-'}</div>
                <div class="p-pts">+{m_top1['points']} แต้ม</div>
            </div>
            <div class="p-card p-silver">
                <div class="p-badge">🥈 อันดับ 2</div>
                <div class="p-team">{m_top2['team_name'] if m_top2 else '-'}</div>
                <div class="p-mgr">{m_top2['player_name'] if m_top2 else '-'}</div>
                <div class="p-pts">+{m_top2['points']} แต้ม</div>
            </div>
            <div class="p-card p-bronze">
                <div class="p-badge">🥉 อันดับ 3</div>
                <div class="p-team">{m_top3['team_name'] if m_top3 else '-'}</div>
                <div class="p-mgr">{m_top3['player_name'] if m_top3 else '-'}</div>
                <div class="p-pts">+{m_top3['points']} แต้ม</div>
            </div>
        </div>
        """

        table_rows = ""
        for r in rows:
            rank = r["rank"]
            rank_class = "gold-row" if rank == 1 else ("silver-row" if rank == 2 else ("bronze-row" if rank == 3 else ""))
            rank_str = f"🥇 1" if rank == 1 else (f"🥈 2" if rank == 2 else (f"🥉 3" if rank == 3 else f"#{rank}"))
            
            table_rows += f"""
            <tr class="{rank_class}">
                <td class="col-center font-bold">{rank_str}</td>
                <td>
                    <div class="t-name">{r['team_name']}</div>
                    <div class="m-name">{r['player_name']}</div>
                </td>
                <td class="col-center font-bold m-pts">+{r['points']}</td>
                <td class="col-center text-muted">{r['season_total']}</td>
            </tr>
            """

        month_sections_html += f"""
        <div class="month-block">
            <div class="month-header">
                <div><h2 class="month-title">📅 {month_title} <span class="gw-badge">({gw_range_str})</span></h2></div>
                <div>{status_tag}</div>
            </div>
            {podium_html}
            <div class="table-wrap">
                <table>
                    <thead>
                        <tr>
                            <th class="col-center" style="width: 70px;">อันดับ</th>
                            <th>ชื่อทีม & ผู้จัดการทีม</th>
                            <th class="col-center" style="width: 130px;">คะแนนเดือนนี้</th>
                            <th class="col-center" style="width: 120px;">คะแนนรวมทั้งซีซั่น</th>
                        </tr>
                    </thead>
                    <tbody>{table_rows}</tbody>
                </table>
            </div>
        </div>
        """

    html = f"""<!DOCTYPE html>
<html lang="th">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Monthly Awards - {league_name}</title>
    <link href="https://fonts.googleapis.com/css2?family=Kanit:wght@300;400;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root {{ --pl-purple: #37003c; --pl-green: #00ff87; --pl-pink: #e90052; --pl-gold: #ffd700; --pl-cyan: #04f5ff; }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ background: linear-gradient(135deg, #14001a 0%, #08000c 100%); font-family: 'Kanit', sans-serif; color: #fff; padding: 25px 15px; display: flex; flex-direction: column; align-items: center; }}
        {ACTION_BAR_CSS}
        .container {{ max-width: 950px; width: 100%; background: rgba(30, 3, 35, 0.95); backdrop-filter: blur(15px); border: 1px solid rgba(255, 255, 255, 0.1); border-radius: 26px; padding: 35px 25px; box-shadow: 0 30px 70px rgba(0,0,0,0.7), 0 0 30px rgba(0, 255, 135, 0.15); }}
        .header {{ text-align: center; margin-bottom: 30px; }}
        .badge-tag {{ display: inline-block; background: linear-gradient(90deg, #00ff87, #04f5ff); color: #120015; padding: 5px 18px; border-radius: 30px; font-size: 12px; font-weight: 800; letter-spacing: 1px; text-transform: uppercase; margin-bottom: 10px; }}
        .main-title {{ font-size: 32px; font-weight: 900; background: linear-gradient(90deg, #ffffff, #00ff87, #04f5ff); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-bottom: 5px; }}
        .sub-title {{ color: #c4a9d4; font-size: 14.5px; }}
        .month-block {{ background: rgba(255, 255, 255, 0.03); border: 1px solid rgba(255, 255, 255, 0.08); border-radius: 20px; padding: 22px; margin-bottom: 30px; }}
        .month-header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 18px; flex-wrap: wrap; gap: 10px; }}
        .month-title {{ font-size: 20px; font-weight: 800; color: #fff; }}
        .gw-badge {{ font-size: 14px; color: var(--pl-green); font-weight: 700; }}
        .m-status {{ padding: 4px 12px; border-radius: 12px; font-size: 11px; font-weight: 700; }}
        .m-status.active {{ background: rgba(0, 255, 135, 0.15); color: var(--pl-green); border: 1px solid rgba(0, 255, 135, 0.3); }}
        .m-status.done {{ background: rgba(255, 255, 255, 0.1); color: #d0c0dd; }}
        .podium-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 12px; margin-bottom: 18px; }}
        .p-card {{ border-radius: 14px; padding: 14px; text-align: center; border: 1px solid rgba(255, 255, 255, 0.1); }}
        .p-gold {{ background: linear-gradient(135deg, rgba(255, 215, 0, 0.15), rgba(255, 215, 0, 0.03)); border-color: rgba(255, 215, 0, 0.5); }}
        .p-silver {{ background: rgba(220, 220, 220, 0.07); border-color: rgba(220, 220, 220, 0.3); }}
        .p-bronze {{ background: rgba(205, 127, 50, 0.07); border-color: rgba(205, 127, 50, 0.3); }}
        .p-badge {{ font-size: 11.5px; font-weight: 800; margin-bottom: 4px; text-transform: uppercase; }}
        .p-gold .p-badge {{ color: var(--pl-gold); }}
        .p-silver .p-badge {{ color: #e0e0e0; }}
        .p-bronze .p-badge {{ color: #cd7f32; }}
        .p-team {{ font-size: 15px; font-weight: 800; color: #fff; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
        .p-mgr {{ font-size: 11.5px; color: #bca5ce; margin-bottom: 6px; }}
        .p-pts {{ font-size: 20px; font-weight: 900; color: var(--pl-green); }}
        .table-wrap {{ overflow-x: auto; border-radius: 14px; border: 1px solid rgba(255, 255, 255, 0.06); }}
        table {{ width: 100%; border-collapse: collapse; text-align: left; }}
        thead {{ background: linear-gradient(90deg, #280030, #17001c); }}
        th {{ padding: 12px 14px; font-size: 12px; font-weight: 700; color: #d1b4e3; text-transform: uppercase; }}
        td {{ padding: 11px 14px; border-bottom: 1px solid rgba(255, 255, 255, 0.04); font-size: 13.5px; }}
        tr:hover {{ background: rgba(255, 255, 255, 0.03); }}
        .col-center {{ text-align: center; }}
        .font-bold {{ font-weight: 800; }}
        .text-muted {{ color: #8f7a9e; }}
        .t-name {{ font-weight: 700; color: #fff; }}
        .m-name {{ font-size: 11px; color: #a992b8; }}
        .m-pts {{ color: var(--pl-green); font-size: 14.5px; }}
        .gold-row {{ background: rgba(255, 215, 0, 0.08); color: var(--pl-gold); }}
        .silver-row {{ background: rgba(220, 220, 220, 0.04); }}
        .bronze-row {{ background: rgba(205, 127, 50, 0.04); }}
        .footer {{ text-align: center; margin-top: 25px; color: #7b628a; font-size: 12px; }}
    </style>
</head>
<body>
    {NAV_ACTION_BAR_HTML}
    <div class="container">
        <div class="header">
            <div class="badge-tag">📅 MONTHLY MANAGER OF THE MONTH | LEAGUE {league_id}</div>
            <h1 class="main-title">MONTHLY AWARDS & LEADERBOARD</h1>
            <p class="sub-title">สรุปคะแนนสะสมและอันดับผู้จัดการทีมยอดเยี่ยมประจำแต่ละเดือน ({league_name})</p>
        </div>
        {month_sections_html}
        <div class="footer">Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')} | FPL Auto Exporter</div>
    </div>
</body>
</html>
"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return output_path

def generate_gw_special_html(league_name, league_id, gw_number, prize_title, is_no_chip_rule, gw_prize_rows, tie_breaker_desc, output_path):
    champ = gw_prize_rows[0] if gw_prize_rows else None
    penalized_users = [r for r in gw_prize_rows if r["deduction"] > 0]
    penalized_users.sort(key=lambda x: x["deduction"], reverse=True)

    var_cards_html = ""
    if penalized_users and is_no_chip_rule:
        for p in penalized_users[:4]:
            orig_rank = p["fpl_rank"]
            new_rank = p["prize_rank"]
            rank_drop = new_rank - orig_rank
            drop_txt = f"📉 ร่วงลง {rank_drop} อันดับ (จาก #{orig_rank} ➔ #{new_rank})" if rank_drop > 0 else f"อันดับ #{new_rank}"
            
            var_cards_html += f"""
            <div class="var-card">
                <div class="var-badge">🚨 โดนริบแต้มชิป</div>
                <div class="var-team">{p['team_name']}</div>
                <div class="var-manager">ผู้จัดการ: {p['player_name']}</div>
                <div class="var-chip-tag">{p['chip']}</div>
                <div class="var-math">
                    <span class="raw-pts">{p['raw_points']} แต้ม</span>
                    <span class="minus-pts">-{p['deduction']} ({p['chip_short']})</span>
                    <span class="equal-sign">=</span>
                    <span class="net-pts">{p['net_points']} แต้มสุทธิ</span>
                </div>
                <div class="var-drop">{drop_txt}</div>
            </div>
            """
    else:
        var_cards_html = f"""
        <div style="grid-column: 1 / -1; text-align: center; color: #00ff87; padding: 20px; font-size: 14px;">
            🎉 ไม่มีผู้จัดการทีมคนใดถูกหักคะแนน หรือสัปดาห์นี้เปิดให้ใช้ชิปได้ตามปกติ!
        </div>
        """

    table_rows = ""
    for r in gw_prize_rows:
        p_rank = r["prize_rank"]
        f_rank = r["fpl_rank"]
        change = f_rank - p_rank

        move_badge = f'<span class="badge-up">▲ +{change}</span>' if change > 0 else (f'<span class="badge-down">▼ {change}</span>' if change < 0 else '<span class="badge-same">▬ 0</span>')
        rank_class = "gold-row" if p_rank == 1 else ("silver-row" if p_rank == 2 else ("bronze-row" if p_rank == 3 else ""))
        rank_display = "🥇 1" if p_rank == 1 else ("🥈 2" if p_rank == 2 else ("🥉 3" if p_rank == 3 else f"#{p_rank}"))

        chip_col = '<span class="chip-none">ไม่ได้ใช้ชิป ✅</span>'
        if r["deduction"] > 0:
            chip_col = f'<span class="chip-alert">⚠️ {r["chip"]} (-{r["deduction"]})</span>'
        elif r["chip"] != "-":
            chip_col = f'<span style="color:#c490ff; font-weight:700;">{r["chip"]}</span>'

        tie_badge = ""
        if r.get("tie_note"):
            tie_badge = f'<div class="tie-note">{r["tie_note"]}</div>'

        table_rows += f"""
        <tr class="{rank_class}">
            <td class="col-center font-bold">{rank_display}</td>
            <td class="col-center text-muted">#{f_rank}</td>
            <td class="col-center">{move_badge}</td>
            <td>
                <div class="team-name">{r['team_name']}</div>
                <div class="mgr-name">{r['player_name']}</div>
                {tie_badge}
            </td>
            <td class="col-center net-col font-bold">{r['net_points']}</td>
            <td class="col-center cap-col">
                <span class="cap-name">© {r['captain_name']}</span>
                <span class="cap-pts">({r['captain_pts']} pts)</span>
            </td>
            <td class="col-center bench-col">{r['bench_pts']}</td>
            <td class="col-center">{chip_col}</td>
        </tr>
        """

    rule_banner = "🚨 สัปดาห์พิเศษ: ห้ามใช้ชิปทุกชนิด" if is_no_chip_rule else "🎉 สัปดาห์แจกรางวัลพิเศษ"

    html = f"""<!DOCTYPE html>
<html lang="th">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>GW{gw_number} Prize Ceremony - {league_name}</title>
    <link href="https://fonts.googleapis.com/css2?family=Kanit:wght@300;400;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root {{ --pl-purple: #37003c; --pl-green: #00ff87; --pl-pink: #e90052; --pl-gold: #ffd700; --pl-cyan: #04f5ff; }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ background: linear-gradient(135deg, #18001c 0%, #08000a 100%); font-family: 'Kanit', sans-serif; color: #fff; padding: 25px 15px; display: flex; flex-direction: column; align-items: center; }}
        {ACTION_BAR_CSS}
        .container {{ max-width: 1000px; width: 100%; background: rgba(30, 2, 34, 0.95); backdrop-filter: blur(20px); border: 1px solid rgba(255, 255, 255, 0.12); border-radius: 28px; padding: 35px 25px; box-shadow: 0 30px 70px rgba(0,0,0,0.7), 0 0 35px rgba(233, 0, 82, 0.2); }}
        .header {{ text-align: center; margin-bottom: 20px; }}
        .badge-rule {{ display: inline-block; background: linear-gradient(90deg, #e90052, #ff5e00); color: white; padding: 6px 20px; border-radius: 30px; font-size: 12.5px; font-weight: 800; letter-spacing: 1.5px; text-transform: uppercase; margin-bottom: 12px; }}
        .main-title {{ font-size: 32px; font-weight: 900; background: linear-gradient(90deg, #ffffff, #ffd700, #00ff87); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-bottom: 6px; }}
        .sub-title {{ color: #d1b4e3; font-size: 14.5px; margin-bottom: 12px; }}
        .tie-rule-box {{ background: rgba(0, 255, 135, 0.08); border: 1px solid rgba(0, 255, 135, 0.25); border-radius: 12px; padding: 8px 16px; font-size: 12.5px; color: #00ff87; display: inline-block; font-weight: 600; }}
        .spotlight-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin: 25px 0; }}
        @media(max-width: 768px) {{ .spotlight-grid {{ grid-template-columns: 1fr; }} }}
        .champ-card {{ background: linear-gradient(145deg, rgba(255, 215, 0, 0.15), rgba(255, 140, 0, 0.05)); border: 2px solid var(--pl-gold); border-radius: 20px; padding: 22px; display: flex; flex-direction: column; justify-content: center; }}
        .crown-icon {{ font-size: 36px; margin-bottom: 6px; }}
        .champ-label {{ font-size: 12px; color: var(--pl-gold); text-transform: uppercase; font-weight: 800; }}
        .champ-team {{ font-size: 22px; font-weight: 800; color: #fff; margin: 4px 0; }}
        .champ-mgr {{ font-size: 13.5px; color: #f0d5a3; margin-bottom: 10px; }}
        .champ-pts {{ font-size: 28px; font-weight: 900; color: var(--pl-green); }}
        .champ-sub {{ font-size: 11.5px; color: #00ff87; }}
        .var-summary-box {{ background: rgba(233, 0, 82, 0.08); border: 2px dashed rgba(233, 0, 82, 0.4); border-radius: 20px; padding: 18px; }}
        .var-box-title {{ font-size: 15px; font-weight: 800; color: var(--pl-pink); margin-bottom: 10px; display: flex; align-items: center; gap: 8px; }}
        .var-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
        @media(max-width: 500px) {{ .var-grid {{ grid-template-columns: 1fr; }} }}
        .var-card {{ background: rgba(0, 0, 0, 0.35); border: 1px solid rgba(233, 0, 82, 0.25); border-radius: 12px; padding: 10px; }}
        .var-badge {{ font-size: 9.5px; font-weight: 800; color: var(--pl-pink); text-transform: uppercase; }}
        .var-team {{ font-size: 13.5px; font-weight: 700; color: #fff; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
        .var-manager {{ font-size: 10.5px; color: #bfa3cf; }}
        .var-chip-tag {{ display: inline-block; background: #963cff; color: #fff; font-size: 9.5px; font-weight: 700; padding: 1px 5px; border-radius: 5px; margin: 3px 0; }}
        .var-math {{ font-size: 10.5px; margin-top: 3px; }}
        .raw-pts {{ color: #a99bb3; text-decoration: line-through; }}
        .minus-pts {{ color: var(--pl-pink); font-weight: 700; }}
        .net-pts {{ color: var(--pl-green); font-weight: 800; }}
        .var-drop {{ font-size: 10.5px; color: #ff99bb; font-weight: 700; margin-top: 4px; }}
        .section-title {{ font-size: 17px; font-weight: 800; color: #fff; margin-bottom: 14px; }}
        .table-wrap {{ overflow-x: auto; border-radius: 16px; border: 1px solid rgba(255, 255, 255, 0.08); }}
        table {{ width: 100%; border-collapse: collapse; text-align: left; }}
        thead {{ background: linear-gradient(90deg, #2c0032, #18001e); }}
        th {{ padding: 12px 10px; font-size: 11.5px; font-weight: 700; color: #d1b4e3; text-transform: uppercase; }}
        td {{ padding: 11px 10px; border-bottom: 1px solid rgba(255, 255, 255, 0.05); font-size: 13px; }}
        tr:hover {{ background: rgba(255, 255, 255, 0.04); }}
        .col-center {{ text-align: center; }}
        .font-bold {{ font-weight: 800; }}
        .text-muted {{ color: #8f7a9e; }}
        .gold-row {{ background: rgba(255, 215, 0, 0.1); color: var(--pl-gold); }}
        .silver-row {{ background: rgba(220, 220, 220, 0.06); }}
        .bronze-row {{ background: rgba(205, 127, 50, 0.06); }}
        .badge-up {{ background: rgba(0, 255, 135, 0.15); color: var(--pl-green); padding: 2px 7px; border-radius: 10px; font-size: 11px; font-weight: 700; }}
        .badge-down {{ background: rgba(233, 0, 82, 0.15); color: var(--pl-pink); padding: 2px 7px; border-radius: 10px; font-size: 11px; font-weight: 700; }}
        .badge-same {{ color: #776082; font-size: 11px; }}
        .team-name {{ font-weight: 700; color: #fff; }}
        .mgr-name {{ font-size: 11px; color: #a992b8; }}
        .chip-none {{ color: #55d697; font-size: 11px; }}
        .chip-alert {{ color: #ff6b8b; font-weight: 700; font-size: 11px; }}
        .net-col {{ color: var(--pl-green); font-size: 15px; }}
        .cap-col {{ font-size: 12px; }}
        .cap-name {{ color: #ffffff; font-weight: 600; display: block; }}
        .cap-pts {{ color: var(--pl-cyan); font-weight: 700; }}
        .bench-col {{ color: #ffd700; font-weight: 700; }}
        .tie-note {{ font-size: 10.5px; color: #ffd700; background: rgba(255, 215, 0, 0.1); padding: 2px 6px; border-radius: 6px; display: inline-block; margin-top: 3px; }}
        .footer {{ text-align: center; margin-top: 20px; color: #7b628a; font-size: 11.5px; }}
    </style>
</head>
<body>
    {NAV_ACTION_BAR_HTML}
    <div class="container">
        <div class="header">
            <div class="badge-rule">{rule_banner}</div>
            <h1 class="main-title">{prize_title}</h1>
            <p class="sub-title">สรุปผลการแจกรางวัลประจำสัปดาห์ <strong>GAMEWEEK {gw_number}</strong> ({league_name})</p>
            <div class="tie-rule-box">📌 {tie_breaker_desc}</div>
        </div>
        <div class="spotlight-grid">
            <div class="champ-card">
                <div class="crown-icon">👑</div>
                <div class="champ-label">🏆 แชมป์ประจำสัปดาห์ (GW{gw_number} WINNER)</div>
                <div class="champ-team">{champ['team_name'] if champ else '-'}</div>
                <div class="champ-mgr">ผู้จัดการทีม: {champ['player_name'] if champ else '-'}</div>
                <div class="champ-pts">{champ['net_points'] if champ else 0} แต้มสุทธิ</div>
                <div class="champ-sub">กัปตัน: © {champ['captain_name']} ({champ['captain_pts']} pts) | Bench: {champ['bench_pts']} pts</div>
            </div>
            <div class="var-summary-box">
                <div class="var-box-title">⚡ บันทึกการตรวจสอบชิป (VAR Audit)</div>
                <div class="var-grid">{var_cards_html}</div>
            </div>
        </div>
        <div class="section-title"><span>📋 ตารางอันดับรับรางวัล GW{gw_number}</span></div>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th class="col-center">อันดับ</th>
                        <th class="col-center">อันดับเดิม</th>
                        <th class="col-center">ขยับ</th>
                        <th>ชื่อทีม & ผู้จัดการทีม</th>
                        <th class="col-center">คะแนนสุทธิ</th>
                        <th class="col-center">กัปตัน (Captain)</th>
                        <th class="col-center">สำรอง (Bench)</th>
                        <th class="col-center">สถานะชิป</th>
                    </tr>
                </thead>
                <tbody>{table_rows}</tbody>
            </table>
        </div>
        <div class="footer">Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')} | FPL Auto Exporter</div>
    </div>
</body>
</html>
"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return output_path

def generate_index_portal_html(league_name, league_id, current_gw, excel_filename, special_gws_list, output_path):
    special_links_html = ""
    for sgw in special_gws_list:
        special_links_html += f"""
        <a href="3_Special_Prize_GW{sgw}.html" class="portal-card prize-card">
            <div class="card-icon">🎁</div>
            <div class="card-title">รางวัลพิเศษ GW{sgw}</div>
            <div class="card-desc">ดูผลการแจกรางวัล & ดราม่าห้อง VAR</div>
        </a>
        """

    html = f"""<!DOCTYPE html>
<html lang="th">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{league_name} - FPL Hub</title>
    <link href="https://fonts.googleapis.com/css2?family=Kanit:wght@300;400;600;700;800;900&display=swap" rel="stylesheet">
    <style>
        :root {{ --pl-purple: #37003c; --pl-green: #00ff87; --pl-pink: #e90052; --pl-cyan: #04f5ff; --pl-gold: #ffd700; }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ background: linear-gradient(135deg, #18001c 0%, #08000a 100%); font-family: 'Kanit', sans-serif; color: #fff; min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 25px 15px; }}
        .hub-container {{ max-width: 800px; width: 100%; background: rgba(30, 2, 34, 0.95); backdrop-filter: blur(20px); border: 1px solid rgba(255, 255, 255, 0.12); border-radius: 28px; padding: 40px 30px; box-shadow: 0 30px 70px rgba(0,0,0,0.7), 0 0 35px rgba(0, 255, 135, 0.15); text-align: center; }}
        .badge-hub {{ display: inline-block; background: linear-gradient(90deg, #00ff87, #04f5ff); color: #120015; padding: 6px 20px; border-radius: 30px; font-size: 13px; font-weight: 800; letter-spacing: 1.5px; text-transform: uppercase; margin-bottom: 15px; }}
        .hub-title {{ font-size: 36px; font-weight: 900; background: linear-gradient(90deg, #ffffff, #00ff87, #04f5ff); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-bottom: 8px; }}
        .hub-sub {{ color: #d1b4e3; font-size: 16px; margin-bottom: 35px; }}
        .grid-cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 16px; margin-bottom: 30px; }}
        .portal-card {{ background: rgba(255, 255, 255, 0.04); border: 1px solid rgba(255, 255, 255, 0.1); border-radius: 20px; padding: 25px 20px; text-decoration: none; color: #fff; display: flex; flex-direction: column; align-items: center; transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); }}
        .portal-card:hover {{ transform: translateY(-5px); border-color: var(--pl-green); box-shadow: 0 10px 30px rgba(0, 255, 135, 0.25); background: rgba(255, 255, 255, 0.07); }}
        .prize-card:hover {{ border-color: var(--pl-pink); box-shadow: 0 10px 30px rgba(233, 0, 82, 0.25); }}
        .card-icon {{ font-size: 42px; margin-bottom: 12px; }}
        .card-title {{ font-size: 18px; font-weight: 800; margin-bottom: 6px; }}
        .card-desc {{ font-size: 12.5px; color: #bba3cc; }}
        .btn-excel {{ display: inline-flex; align-items: center; gap: 8px; background: rgba(255, 255, 255, 0.08); border: 1px solid rgba(255, 255, 255, 0.2); color: #fff; padding: 12px 24px; border-radius: 30px; font-size: 14px; font-weight: 700; text-decoration: none; transition: 0.2s; }}
        .btn-excel:hover {{ background: #1f4e79; border-color: #2a6db0; }}
        .footer {{ margin-top: 30px; color: #7b628a; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="hub-container">
        <div class="badge-hub">LEAGUE ID: {league_id}</div>
        <h1 class="hub-title">{league_name}</h1>
        <p class="hub-sub">ศูนย์รวมตารางคะแนนและสรุปผลรางวัลประจำฤดูกาล (อัปเดตล่าสุด: GW {current_gw})</p>

        <div class="grid-cards">
            <a href="1_Weekly_Standings_GW{current_gw}.html" class="portal-card">
                <div class="card-icon">📊</div>
                <div class="card-title">ตารางคะแนนสัปดาห์นี้</div>
                <div class="card-desc">ดูอันดับล่าสุด, แต้มรวม, กัปตัน และการขยับอันดับ</div>
            </a>
            <a href="2_Monthly_Awards_GW{current_gw}.html" class="portal-card">
                <div class="card-icon">📅</div>
                <div class="card-title">สรุปผลงานประจำเดือน</div>
                <div class="card-desc">ดูแชมป์ประจำเดือน (Manager of the Month)</div>
            </a>
            {special_links_html}
        </div>

        <div>
            <a href="{excel_filename}" download class="btn-excel">
                <span>📥 ดาวน์โหลดไฟล์ Excel รวมทั้งฤดูกาล (.xlsx)</span>
            </a>
        </div>

        <div class="footer">Fantasy Premier League Automated Dashboard | Updated on {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
    </div>
</body>
</html>
"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    return output_path

def fetch_and_export_full_season():
    config = load_config()
    league_id = config.get("league_id", 506845)
    no_chip_gws = config.get("rules", {}).get("no_chip_gameweeks", [1])
    special_prizes = config.get("prizes", {}).get("special_weekly_prizes", {})
    monthly_places = config.get("prizes", {}).get("monthly_top_places", 3)
    tie_breaker_desc = config.get("tie_breaker", {}).get("description", "กรณีแต้ม GW เท่ากัน: 1) คะแนนกัปตันสูงกว่า 2) Bench รวมสูงกว่า 3) จับฉลาก")

    print(f"==================================================")
    print(f"  FPL Dynamic Exporter & Graphic Studio (League {league_id})")
    print(f"==================================================")
    
    events, current_gw, player_map, months_gw_map, gw_to_month = get_bootstrap_data()
    
    # โหลด live points สำหรับทุกสัปดาห์ที่ต้องคำนวณไทเบรกเกอร์/รางวัลพิเศษ + สัปดาห์ปัจจุบัน
    target_fetch_gws = set(no_chip_gws)
    for g_str in special_prizes.keys():
        if int(g_str) <= current_gw:
            target_fetch_gws.add(int(g_str))
    target_fetch_gws.add(current_gw)

    live_points_cache = {}
    for g in target_fetch_gws:
        if g <= current_gw:
            live_points_cache[g] = get_live_points_for_gw(g)

    print(f"👉 สัปดาห์ปัจจุบัน: GW {current_gw}")
    print(f"📅 พบการแบ่งเดือนทั้งหมด: {len(months_gw_map)} เดือน\n")

    print(f"กำลังดึงรายชื่อสมาชิกใน League ID: {league_id}...")
    league_name = "FPL_League"
    members = []
    page = 1
    while True:
        url = f"https://fantasy.premierleague.com/api/leagues-classic/{league_id}/standings/?page_standings={page}"
        res = requests.get(url, headers=HEADERS, verify=False)
        if res.status_code != 200:
            break
        data = res.json()
        if "league" in data:
            league_name = data["league"].get("name", league_name)
        results = data.get("standings", {}).get("results", [])
        if not results:
            break
        members.extend(results)
        if not data.get("standings", {}).get("has_next"):
            break
        page += 1

    print(f"✅ พบสมาชิกทั้งหมด {len(members)} ทีม กำลังคำนวณข้อมูลกราฟฟิกและ Excel...\n")

    season_rows = []
    chips_tracker_rows = []
    display_weekly_graphic_rows = []
    
    special_gws_data = {int(gw): [] for gw in special_prizes.keys() if int(gw) <= current_gw}
    for g in no_chip_gws:
        if g <= current_gw and g not in special_gws_data:
            special_gws_data[g] = []

    monthly_data_map = {}
    for m_title, gws in months_gw_map.items():
        monthly_data_map[m_title] = {"gws": gws, "rows": []}

    for idx, m in enumerate(members, 1):
        entry_id = m.get("entry")
        team_name = m.get("entry_name")
        player_name = m.get("player_name")
        fpl_rank = m.get("rank")
        last_rank = m.get("last_rank")
        rank_change = (last_rank - fpl_rank) if (last_rank and fpl_rank) else 0

        hist_url = f"https://fantasy.premierleague.com/api/entry/{entry_id}/history/"
        h_res = requests.get(hist_url, headers=HEADERS, verify=False)
        
        gw_official_points = {}
        used_chips_dict = {}
        
        if h_res.status_code == 200:
            h_data = h_res.json()
            for cur in h_data.get("current", []):
                gw = cur.get("event")
                pts = cur.get("points", 0) - cur.get("event_transfers_cost", 0)
                gw_official_points[gw] = pts

            for c in h_data.get("chips", []):
                used_chips_dict[c.get("event")] = c.get("name")

        official_total_pts = sum(gw_official_points.get(g, 0) for g in range(1, current_gw + 1))
        
        # ดึงข้อมูล Picks ของสัปดาห์ปัจจุบัน เพื่อหากัปตัน และตัวสำรอง
        cur_captain_name = "-"
        cur_captain_pts = 0
        cur_bench_pts = 0
        
        cur_pick_url = f"https://fantasy.premierleague.com/api/entry/{entry_id}/event/{current_gw}/picks/"
        cur_pick_res = requests.get(cur_pick_url, headers=HEADERS, verify=False)
        if cur_pick_res.status_code == 200:
            p_data_cur = cur_pick_res.json()
            gw_live_cur = live_points_cache.get(current_gw, {})
            for pick in p_data_cur.get("picks", []):
                p_id = pick.get("element")
                p_pts = gw_live_cur.get(p_id, 0)
                mult = pick.get("multiplier", 1)
                if pick.get("is_captain"):
                    cur_captain_name = player_map.get(p_id, "Captain")
                    cur_captain_pts = p_pts * (mult if mult > 0 else 1)
                if pick.get("position", 1) >= 12:
                    cur_bench_pts += p_pts

        season_row = {
            "ชื่อทีม (Team Name)": team_name,
            "ผู้จัดการทีม (Manager)": player_name,
            "คะแนนรวม (Total Points)": official_total_pts,
            f"กัปตัน GW{current_gw} (Captain)": f"{cur_captain_name} ({cur_captain_pts} pts)",
            f"สำรอง GW{current_gw} (Bench)": cur_bench_pts,
        }
        for g in range(1, 39):
            season_row[f"GW{g}"] = gw_official_points.get(g, "") if g <= current_gw else ""
        season_rows.append(season_row)

        # คำนวณรางวัลพิเศษ & Tie-breaker สำหรับสัปดาห์รางวัล
        for target_gw in special_gws_data.keys():
            raw_pts = gw_official_points.get(target_gw, 0)
            chip_used = used_chips_dict.get(target_gw)
            deduction = 0
            chip_short = ""
            reason = "เล่นตามปกติ (ผ่านเกณฑ์)"
            
            captain_name = "-"
            captain_pts = 0
            bench_pts = 0

            # ดึง picks เพื่อหาแต้มกัปตัน และแต้มตัวสำรอง (Tie-breaker)
            p_url = f"https://fantasy.premierleague.com/api/entry/{entry_id}/event/{target_gw}/picks/"
            p_res = requests.get(p_url, headers=HEADERS, verify=False)
            if p_res.status_code == 200:
                p_data = p_res.json()
                gw_live = live_points_cache.get(target_gw, {})
                bench_names = []
                
                for pick in p_data.get("picks", []):
                    p_id = pick.get("element")
                    p_pts = gw_live.get(p_id, 0)
                    mult = pick.get("multiplier", 1)

                    if pick.get("is_captain"):
                        captain_name = player_map.get(p_id, "Captain")
                        captain_pts = p_pts * (mult if mult > 0 else 1)

                    if pick.get("position", 1) >= 12:
                        bench_pts += p_pts
                        bench_names.append(f"{player_map.get(p_id, '')}({p_pts})")

                if target_gw in no_chip_gws and chip_used:
                    if chip_used == "bboost":
                        chip_short = "BB"
                        deduction = bench_pts
                        reason = f"ใช้ BB -> หักสำรอง 4 คน [{', '.join(bench_names)}]"
                    elif chip_used == "3xc":
                        chip_short = "TC"
                        base_c_pts = gw_live.get(pick.get("element"), 0)
                        for pick in p_data.get("picks", []):
                            if pick.get("is_captain"):
                                base_c_pts = gw_live.get(pick.get("element"), 0)
                                break
                        deduction = base_c_pts
                        reason = f"ใช้ TC -> หักกัปตัน {captain_name} ({base_c_pts} pts)"
                    elif chip_used in ["wildcard", "freehit"]:
                        chip_short = chip_used.upper()
                        reason = f"ใช้ {CHIP_DISPLAY.get(chip_used, chip_used)}"

            net_prize_pts = raw_pts - deduction
            special_gws_data[target_gw].append({
                "fpl_rank": fpl_rank,
                "team_name": team_name,
                "player_name": player_name,
                "raw_points": raw_pts,
                "chip": CHIP_DISPLAY.get(chip_used, "-") if chip_used else "-",
                "chip_short": chip_short,
                "deduction": deduction,
                "net_points": net_prize_pts,
                "captain_name": captain_name,
                "captain_pts": captain_pts,
                "bench_pts": bench_pts,
                "reason": reason
            })

        for m_title, m_info in monthly_data_map.items():
            played_gws = [g for g in m_info["gws"] if g <= current_gw]
            if played_gws:
                m_score = sum(gw_official_points.get(g, 0) for g in played_gws)
                m_info["rows"].append({
                    "team_name": team_name,
                    "player_name": player_name,
                    "points": m_score,
                    "season_total": official_total_pts
                })

        chips_tracker_rows.append({
            "ชื่อทีม (Team Name)": team_name,
            "ผู้จัดการทีม (Manager)": player_name,
            "Wildcard": next((f"GW{ev}" for ev, name in used_chips_dict.items() if name == "wildcard"), "-"),
            "Triple Captain": next((f"GW{ev}" for ev, name in used_chips_dict.items() if name == "3xc"), "-"),
            "Bench Boost": next((f"GW{ev}" for ev, name in used_chips_dict.items() if name == "bboost"), "-"),
            "Free Hit": next((f"GW{ev}" for ev, name in used_chips_dict.items() if name == "freehit"), "-"),
        })

        current_gw_chip = used_chips_dict.get(current_gw)
        display_weekly_graphic_rows.append({
            "rank": fpl_rank,
            "rank_change": rank_change,
            "team_name": team_name,
            "player_name": player_name,
            "gw_pts": gw_official_points.get(current_gw, 0),
            "total_pts": official_total_pts,
            "chip": CHIP_DISPLAY.get(current_gw_chip, "-") if current_gw_chip else "-",
            "captain_name": cur_captain_name,
            "captain_pts": cur_captain_pts,
            "bench_pts": cur_bench_pts
        })

        print(f"[{idx}/{len(members)}] {team_name} | แต้มรวม: {official_total_pts} | GW{current_gw}: {gw_official_points.get(current_gw, 0)} (© {cur_captain_name}: {cur_captain_pts} pts)")
        time.sleep(0.05)

    season_rows.sort(key=lambda x: x["คะแนนรวม (Total Points)"], reverse=True)
    for rank, row in enumerate(season_rows, 1):
        row_with_rank = {"อันดับ (Rank)": rank}
        row_with_rank.update(row)
        season_rows[rank - 1] = row_with_rank

    # การเรียงอันดับรางวัลพิเศษตามกฎ Tie-breaker: 1) แต้มสุทธิ 2) แต้มกัปตัน 3) แต้มสำรอง
    special_excel_dfs = {}
    for target_gw, rows in special_gws_data.items():
        rows.sort(key=lambda x: (x["net_points"], x["captain_pts"], x["bench_pts"]), reverse=True)
        gw_excel_list = []
        
        for rank, row in enumerate(rows, 1):
            row["prize_rank"] = rank
            
            # ตรวจสอบการเสมอ (Tie-breaker check)
            tie_note = ""
            if rank > 1:
                prev = rows[rank - 2]
                if row["net_points"] == prev["net_points"]:
                    if row["captain_pts"] < prev["captain_pts"]:
                        tie_note = f"แพ้ไทเบรกเกอร์กัปตัน ({row['captain_pts']} vs {prev['captain_pts']})"
                    elif row["bench_pts"] < prev["bench_pts"]:
                        tie_note = f"แพ้ไทเบรกเกอร์สำรอง ({row['bench_pts']} vs {prev['bench_pts']})"
                    else:
                        tie_note = "🎲 แต้มเท่ากันทุกเกณฑ์ (ต้องจับฉลาก)"
                        prev["tie_note"] = "🎲 แต้มเท่ากันทุกเกณฑ์ (ต้องจับฉลาก)"

            row["tie_note"] = tie_note

            gw_excel_list.append({
                f"อันดับรางวัล GW{target_gw}": rank,
                "อันดับเดิม FPL": row["fpl_rank"],
                "ชื่อทีม (Team Name)": row["team_name"],
                "ผู้จัดการทีม (Manager)": row["player_name"],
                f"คะแนนสุทธิ GW{target_gw}": row["net_points"],
                "กัปตัน": f"© {row['captain_name']} ({row['captain_pts']} pts)",
                "แต้มตัวสำรอง (Bench)": row["bench_pts"],
                "คะแนน FPL ทางการ": row["raw_points"],
                f"ชิปที่ใช้ใน GW{target_gw}": row["chip"],
                "แต้มที่ถูกตัดออก": f"-{row['deduction']}" if row["deduction"] > 0 else "0",
                "กฎตัดสิน / หมายเหตุ": tie_note if tie_note else row["reason"]
            })
        import pandas as pd
        sheet_title = f"🎁 รางวัล GW{target_gw}"
        special_excel_dfs[sheet_title] = pd.DataFrame(gw_excel_list)

    for m_title, m_info in monthly_data_map.items():
        m_info["rows"].sort(key=lambda x: x["points"], reverse=True)
        for r_idx, r in enumerate(m_info["rows"], 1):
            r["rank"] = r_idx

    display_weekly_graphic_rows.sort(key=lambda x: x["rank"])

    monthly_excel_rows = []
    for s_row in season_rows:
        t_name = s_row["ชื่อทีม (Team Name)"]
        m_name_mgr = s_row["ผู้จัดการทีม (Manager)"]
        m_row = {
            "อันดับรวม": s_row["อันดับ (Rank)"],
            "ชื่อทีม (Team Name)": t_name,
            "ผู้จัดการทีม (Manager)": m_name_mgr,
            "คะแนนรวม": s_row["คะแนนรวม (Total Points)"]
        }
        for month_title, gw_list in months_gw_map.items():
            played_gws = [g for g in gw_list if g <= current_gw]
            if played_gws:
                m_total = sum(s_row.get(f"GW{g}", 0) for g in played_gws)
                m_row[f"{month_title} (GW{min(gw_list)}-{max(gw_list)})"] = m_total
            else:
                m_row[f"{month_title} (GW{min(gw_list)}-{max(gw_list)})"] = "-"
        monthly_excel_rows.append(m_row)

    import pandas as pd
    sheets = {
        "🏆 ตารางคะแนนรวม (GW1-38)": pd.DataFrame(season_rows)
    }
    sheets.update(special_excel_dfs)
    sheets["📅 สรุปรายเดือน (Monthly)"] = pd.DataFrame(monthly_excel_rows)
    sheets["⚡ ประวัติการใช้ชิป (Chips)"] = pd.DataFrame(chips_tracker_rows)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    today_str = datetime.now().strftime("%Y-%m-%d")
    
    # 1. โฟลเดอร์ Archive
    output_dir = os.path.join(script_dir, "FPL", today_str)
    os.makedirs(output_dir, exist_ok=True)

    # 2. โฟลเดอร์ Web Root
    web_dir = os.path.join(script_dir, "FPL_Web")
    os.makedirs(web_dir, exist_ok=True)

    base_name = f"FPL_League_{league_id}_GW{current_gw}"
    
    # บันทึกไฟล์ Excel
    excel_path = os.path.join(output_dir, f"{base_name}_Full_Season.xlsx")
    saved_excel = save_excel_safe(sheets, excel_path)
    
    # กราฟฟิก 1: ตารางคะแนนสัปดาห์ (พร้อมกัปตัน & สำรอง)
    weekly_html_path = os.path.join(output_dir, f"1_Weekly_Standings_GW{current_gw}.html")
    generate_weekly_standings_html(league_name, league_id, current_gw, display_weekly_graphic_rows, weekly_html_path)

    # กราฟฟิก 2: ประจำเดือน
    monthly_html_path = os.path.join(output_dir, f"2_Monthly_Awards_GW{current_gw}.html")
    generate_monthly_awards_html(league_name, league_id, current_gw, monthly_data_map, monthly_places, monthly_html_path)

    # กราฟฟิก 3: รางวัลพิเศษ (พร้อมกัปตัน & สำรอง)
    prize_title_current = special_prizes.get(str(current_gw), f"🎁 รางวัลประจำสัปดาห์ GW{current_gw}")
    is_no_chip_current = current_gw in no_chip_gws
    special_prize_html_path = None

    if current_gw in special_gws_data:
        special_prize_html_path = os.path.join(output_dir, f"3_Special_Prize_GW{current_gw}.html")
        generate_gw_special_html(league_name, league_id, current_gw, prize_title_current, is_no_chip_current, special_gws_data[current_gw], tie_breaker_desc, special_prize_html_path)

    # สร้างหน้า Portal Hub (index.html)
    index_html_path = os.path.join(output_dir, "index.html")
    generate_index_portal_html(league_name, league_id, current_gw, os.path.basename(saved_excel), list(special_gws_data.keys()), index_html_path)

    # ซิงค์ไฟล์ทั้งหมดไปยัง FPL_Web
    for f_name in os.listdir(output_dir):
        s_file = os.path.join(output_dir, f_name)
        d_file = os.path.join(web_dir, f_name)
        if os.path.isfile(s_file):
            shutil.copy2(s_file, d_file)

    print(f"\n==================================================")
    print(f"🎉 นำข้อมูลกัปตัน & สำรองขึ้น Dashboard และ Excel ครบทุกจุดเรียบร้อย!")
    print(f"==================================================")

if __name__ == "__main__":
    fetch_and_export_full_season()
